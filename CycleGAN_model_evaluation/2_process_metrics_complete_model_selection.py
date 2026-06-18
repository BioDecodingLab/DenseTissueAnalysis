"""
Script para calcular promedio y SD de las 4 muestras (Ch2, Ch3, G1, G2)
por modelo/método, y marcar en negrita los 3 mejores valores por métrica.
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import sys
import os

# ─── CONFIGURACIÓN ────────────────────────────────────────────────────────────
LOWER_IS_BETTER  = {"MSE", "MAE", "LPIPS"}
HIGHER_IS_BETTER = {"PSNR", "SSIM3D", "PCC", "SII"}

METRICS = ["MSE", "MAE", "PSNR", "SSIM3D", "PCC", "LPIPS", "SII"]
SAMPLES = ["Ch2", "Ch3", "G1", "G2"]
# ──────────────────────────────────────────────────────────────────────────────


def identify_sample(image_name: str) -> str | None:
    for s in SAMPLES:
        if s.lower() in image_name.lower():
            return s
    return None


def process_file(input_path: str, output_path: str | None = None):
    if output_path is None:
        base, _ = os.path.splitext(input_path)
        output_path = f"{base}_summary.xlsx"

    # ── Leer datos ────────────────────────────────────────────────────────────
    ext = os.path.splitext(input_path)[1].lower()
    df = pd.read_csv(input_path) if ext == ".csv" else pd.read_excel(input_path)

    df.columns = df.columns.str.strip()
    df["sample"] = df["image"].astype(str).apply(identify_sample)
    df = df[df["method"].astype(str).str.strip().str.lower() != "raw"]
    df = df.dropna(subset=["sample"])
    df["method"] = pd.to_numeric(df["method"], errors="coerce")
    df = df.dropna(subset=["method"])

    # ── Calcular promedio y SD por método ────────────────────────────────────
    grouped  = df.groupby("method")[METRICS]
    mean_df  = grouped.mean().reset_index()
    std_df   = grouped.std().reset_index()

    # ── Construir tabla: una fila por método, mean y SD en columnas pares ────
    rows = []
    for _, mrow in mean_df.iterrows():
        method = int(mrow["method"])
        row = {"method": method}
        srow = std_df.loc[std_df["method"] == mrow["method"]].iloc[0]
        for m in METRICS:
            row[f"{m}_mean"] = mrow[m]
            row[f"{m}_SD"]   = srow[m]
        rows.append(row)

    result_df = pd.DataFrame(rows)

    # Orden de columnas: method, luego pares mean/SD por métrica
    ordered_cols = ["method"]
    for m in METRICS:
        ordered_cols += [f"{m}_mean", f"{m}_SD"]
    result_df = result_df[ordered_cols]

    # ── Exportar a Excel ──────────────────────────────────────────────────────
    result_df.to_excel(output_path, index=False, sheet_name="Summary")

    # ── Formatear con openpyxl ────────────────────────────────────────────────
    wb = load_workbook(output_path)
    ws = wb["Summary"]

    base_font   = Font(name="Arial", size=10)
    header_font = Font(name="Arial", size=10, bold=True)
    mean_fill   = PatternFill("solid", start_color="DDEEFF")  # azul claro → mean
    sd_fill     = PatternFill("solid", start_color="EEF0EE")  # gris claro  → SD

    col_map = {col: idx + 1 for idx, col in enumerate(result_df.columns)}

    # Encabezados
    for col_idx, col_name in enumerate(result_df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font      = header_font
        cell.fill      = mean_fill if col_name.endswith("_mean") or col_name == "method" else sd_fill
        cell.alignment = Alignment(horizontal="center")

    # Datos
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            col_name = result_df.columns[cell.column - 1]
            cell.font = base_font
            if col_name.endswith("_SD"):
                cell.fill = PatternFill("solid", start_color="F5F5F5")
            if isinstance(cell.value, float):
                cell.number_format = "0.00000"
                cell.alignment     = Alignment(horizontal="right")

    # ── Marcar los 3 mejores valores de cada métrica (columna _mean) ──────────
    for metric in METRICS:
        col_name = f"{metric}_mean"
        if col_name not in col_map:
            continue
        col_idx = col_map[col_name]

        values = result_df[col_name].values.astype(float)
        valid_mask   = ~np.isnan(values)
        valid_values = values[valid_mask]
        valid_rows   = np.where(valid_mask)[0]   # índices base-0 en result_df

        if len(valid_values) == 0:
            continue

        if metric in LOWER_IS_BETTER:
            sorted_idx = np.argsort(valid_values)
        else:
            sorted_idx = np.argsort(valid_values)[::-1]

        top3 = valid_rows[sorted_idx[:3]]

        for df_idx in top3:
            excel_row = int(df_idx) + 2   # +1 header, +1 base-0→base-1
            cell = ws.cell(row=excel_row, column=col_idx)
            cell.font = Font(name="Arial", size=10, bold=True, color="1F497D")


    # ── ANÁLISIS MULTIMÉTRICA SEGÚN GUÍA ──────────────────────────────────
    higher_better = {m: (m in HIGHER_IS_BETTER) for m in METRICS}

    metric_df = mean_df.set_index("method")[METRICS].copy()

    # Normalización min-max orientada
    norm = pd.DataFrame(index=metric_df.index)
    for m in METRICS:
        x = metric_df[m]
        rng = x.max() - x.min()
        if rng == 0:
            norm[m] = 1.0
        elif higher_better[m]:
            norm[m] = (x - x.min()) / rng
        else:
            norm[m] = (x.max() - x) / rng

    fidelidad = ["MSE", "MAE", "PSNR", "PCC"]
    perceptual = ["SSIM3D", "LPIPS", "SII"]

    F_fid = norm[fidelidad].mean(axis=1)
    F_per = norm[perceptual].mean(axis=1)

    w_fid, w_per = 0.30, 0.70
    score = w_fid * F_fid + w_per * F_per

    # Borda
    borda = pd.Series(0.0, index=metric_df.index)
    n = len(metric_df)
    for m in METRICS:
        borda += (n - metric_df[m].rank(ascending=not higher_better[m]))

    # TOPSIS
    V = norm.values / norm.shape[1]
    db = np.sqrt(((V - V.max(axis=0)) ** 2).sum(axis=1))
    dw = np.sqrt(((V - V.min(axis=0)) ** 2).sum(axis=1))
    topsis = dw / (db + dw)

    # Pareto
    M = metric_df.copy()
    for m in METRICS:
        if not higher_better[m]:
            M[m] = -M[m]

    pareto = []
    for i in M.index:
        dominated = any(
            ((M.loc[j] >= M.loc[i]).all() and (M.loc[j] > M.loc[i]).any())
            for j in M.index if j != i
        )
        pareto.append(not dominated)

    # Hoja 1: detalle completo
    analysis_df = norm.copy()
    analysis_df["Fidelity_score"] = F_fid
    analysis_df["Perceptual_score"] = F_per
    analysis_df["Consolidated_score"] = score
    analysis_df["Borda_score"] = borda
    analysis_df["TOPSIS_score"] = topsis
    analysis_df["Pareto_optimal"] = pareto
    analysis_df = analysis_df.reset_index()

    # Hoja 2: ranking
    ranking_df = pd.DataFrame({
        "method": metric_df.index,
        "Fidelity_score": F_fid,
        "Perceptual_score": F_per,
        "Consolidated_score": score,
        "Borda_score": borda,
        "TOPSIS_score": topsis,
        "Pareto_optimal": pareto
    }).sort_values("Consolidated_score", ascending=False)

    ranking_df.insert(0, "Rank", range(1, len(ranking_df) + 1))

    # Hoja 3: sensibilidad
    sens_rows = []
    for wp in [0.0, 0.25, 0.50, 0.70, 1.0]:
        s = (1 - wp) * F_fid + wp * F_per
        ordered = s.sort_values(ascending=False)

        sens_rows.append({
            "Perceptual_weight": wp,
            "Fidelity_weight": 1 - wp,
            "Winner": ordered.index[0],
            "Winner_score": ordered.iloc[0],
            "Second": ordered.index[1] if len(ordered) > 1 else None,
            "Third": ordered.index[2] if len(ordered) > 2 else None,
        })

    sensitivity_df = pd.DataFrame(sens_rows)

    with pd.ExcelWriter(output_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        analysis_df.to_excel(writer, sheet_name="Model_Selection", index=False)
        ranking_df.to_excel(writer, sheet_name="Ranking", index=False)
        sensitivity_df.to_excel(writer, sheet_name="Sensitivity_Analysis", index=False)

    wb = load_workbook(output_path)

    for sheet_name in ["Model_Selection", "Ranking", "Sensitivity_Analysis"]:
        wsx = wb[sheet_name]

        for row in wsx.iter_rows():
            for cell in row:
                if isinstance(cell.value, float):
                    cell.number_format = "0.00000"

        for col in wsx.columns:
            width = max(len(str(c.value)) if c.value is not None else 0 for c in col)
            wsx.column_dimensions[get_column_letter(col[0].column)].width = max(12, width + 2)

    rank_ws = wb["Ranking"]
    winner_row = 2
    for cell in rank_ws[winner_row]:
        cell.font = Font(name="Arial", size=10, bold=True)
        cell.fill = PatternFill("solid", start_color="C6EFCE")

    wb.save(output_path)


    # Ajustar ancho de columnas
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(10, max_len + 2)

    wb.save(output_path)
    print(f"✅ Archivo guardado en: {output_path}")
    return output_path


# ─── PUNTO DE ENTRADA ─────────────────────────────────────────────────────────
if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python process_metrics.py <archivo.xlsx/csv> [salida.xlsx]")
        sys.exit(1)

    input_file  = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else None
    process_file(input_file, output_file)
